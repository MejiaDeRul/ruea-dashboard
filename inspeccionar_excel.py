# inspeccionar_excel.py
# Uso: python inspeccionar_excel.py "D:/ruta/MACRO MADRE SDR.xlsx"

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as e:
    print("Necesitas instalar openpyxl:  pip install openpyxl")
    sys.exit(1)

def normalize(v):
    if v is None:
        return ""
    return str(v).strip()

def guess_header_row(ws, max_scan_rows=15):
    # Busca la primera fila con >= 2 celdas no vacías (posible encabezado)
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        vals = [normalize(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        if sum(v != "" for v in vals) >= 2:
            return r
    return 1

def read_headers(ws, header_row):
    cols = [normalize(ws.cell(row=header_row, column=c).value) for c in range(1, ws.max_column + 1)]
    # recortar vacíos al final
    while cols and cols[-1] == "":
        cols.pop()
    return cols

def guess_module(sheet_name, headers):
    s = sheet_name.lower()
    cols = " ".join(h.lower() for h in headers)
    if "ruea" in s: return "ruea"
    if "comer" in s or "comercial" in s: return "comercializacion"
    if "indic" in s: return "indicadores"
    if "nodo" in s: return "nodos"
    if ("anio" in cols and "estrategia" in cols) or "monto" in cols: return "comercializacion"
    if "cumplimiento" in cols or "indicador" in cols: return "indicadores"
    if any(k in cols for k in ["corregimiento", "vereda", "linea_productiva", "escolaridad"]): return "ruea"
    if any(k in cols for k in ["id_nodo", "nombre_nodo", "productor_id"]): return "nodos"
    return ""

def main(xlsx_path:str):
    p = Path(xlsx_path)
    if not p.exists():
        print("No se encontró el archivo:", p)
        sys.exit(1)

    wb = load_workbook(p, read_only=True, data_only=True)
    print(f"Archivo: {p}")
    print("Hojas:", [ws.title for ws in wb.worksheets])
    print("-" * 80)

    out_lines = ["hoja,fila_encabezado,filas_aprox,num_columnas,modulo_probable,primeras_columnas"]
    for ws in wb.worksheets:
        try:
            header_row = guess_header_row(ws)
            headers = read_headers(ws, header_row)
            n_rows_est = max(ws.max_row - header_row, 0)
            modulo = guess_module(ws.title, headers)
            primeras = "; ".join(headers[:25])
            print(f"[{ws.title}] encabezado≈fila {header_row} | filas≈{n_rows_est} | cols={len(headers)} | módulo≈{modulo}")
            print("  columnas:", primeras)
            print("-" * 80)
            # CSV
            out_lines.append(f"{ws.title},{header_row},{n_rows_est},{len(headers)},{modulo},\"{primeras}\"")
        except Exception as e:
            print(f"[{ws.title}] Error leyendo: {e}")
            print("-" * 80)

    # Guardar un CSV resumen al lado del Excel
    csv_path = p.with_name(p.stem + "_resumen_hojas.csv")
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines))
    print("Resumen guardado en:", csv_path)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print('Uso: python inspeccionar_excel.py "D:/ruta/MACRO MADRE SDR.xlsx"')
        sys.exit(1)
    main(sys.argv[1])
